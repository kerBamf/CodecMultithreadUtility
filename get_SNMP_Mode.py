from os import environ
from dotenv import load_dotenv
from get_info import get_info
import xml.etree.ElementTree as ET
from Utils.logger import log_info
import openpyxl
import xml.etree.ElementTree as ET
from dataclasses import dataclass
import concurrent.futures

class CustomException(Exception):
    pass

#Loading environment variable
load_dotenv()
PASSCODE = environ.get('PASSCODE')
LOGPATH = environ.get('LOGPATH')
headers = {
    'Authorization': f'basic {PASSCODE}'
}

#Setting up logger
def message(string, codec, path=LOGPATH):
    log_info(string, codec, path)
    print(string)

@dataclass
class Codec:
    name: str
    ip: str
    snmp_mode: str

def get_SNMP_Mode(codec):
    try:
        info_dict = get_info(codec.ip, 'Configuration/NetworkServices/SNMP/Mode')
        xml_root = ET.fromstring(info_dict['response'])
        mode = xml_root.find('.//Mode').text
        message(f'{codec.name} SNMP Mode: {mode}', codec.name)
        codec.mode = mode
    except Exception:
        codec.mode = 'Error Retrieving Info'
    return codec
    

def build_excel():
    #Loading excel workbook
    wb = openpyxl.load_workbook('../excel_files/ExpresswayRegCodecs.xlsx')
    ws = wb.active
    codecs_processed = [['Name', 'IP', 'SNMP Mode']]

    # for value in ws.iter_rows(min_row=2, min_col=2, max_col=3, values_only=True):
    #     codec = Codec(value[0], value[1], '')
    #     try:
    #         codec.mode = get_SNMP_Mode(codec)
    #         codecs_processed.append([codec.name, codec.ip, codec.mode])
    #     except Exception:
    #         codecs_processed.append([codec.name, codec.ip, 'Error getting info'])
    #         continue

    #Converting Excel data into usable dictionaries
    codec_list = []
    for value in ws.iter_rows(min_row=2, min_col=2, max_col=3, values_only=True):
        codec = Codec(value[0], value[1], '')
        codec_list.append(codec)

    #Feeding through multithreader to obtain data from codecs
    with concurrent.futures.ThreadPoolExecutor() as executor:
        futures = {executor.submit(get_SNMP_Mode, codec): codec for codec in codec_list}
    
    for future in concurrent.futures.as_completed(futures):
        codec = future.result()
        codecs_processed.append([codec.name, codec.ip, codec.mode])


    #Creating new Excel file
    new_wb = openpyxl.Workbook()
    new_ws = new_wb.active
    new_ws.title = "Codec SNMP Modes"
    ft = openpyxl.styles.Font(bold=True)

    for row in codecs_processed:
        new_ws.append(row)

    for row in new_ws["A1:C1"]:
        for cell in row:
            cell.font = ft

    new_wb.save('../excel_files/CodecSNMPModes.xlsx')


if __name__ == '__main__':
    build_excel()