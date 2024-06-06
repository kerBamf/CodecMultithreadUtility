from os import listdir, getcwd
from os.path import isfile, join
from openpyxl import load_workbook

class Custom_Exception(Exception):
    pass

#Finding relative path
path = f'{getcwd()}/../excel_files'

# Finding files
def find_files():
    file_list = [file for file in listdir(path) if isfile(join(path, file))]
    list_dict = {}
    option_string = ''
    for idx, file in enumerate(file_list):
        list_dict.update({idx: file})
        option_string = option_string + f'\t{file}: {idx}\r\n'

    output = {
        'files': list_dict,
        'options': option_string
    }

    return output

found_files = find_files()
def user_choice():
    choice = int(input(f'Select which excel file to use:\r\n{found_files['options']}\r\nFile number selection: '))
    for key in list(found_files['files'].keys()):
        if choice == key:
            if 'y' == input(f'\n\rYou have selected {found_files['files'][key]}. Proceed? (y/n): '):
                return choice
            else:
                print('Selection Cancelled')
                return user_choice()
    print('Your selection is invalid.')
    return user_choice()

selected_file = found_files['files'][int(user_choice())]



def excel_parser():
    wb = load_workbook(f'{path}/{selected_file}')
    ws = wb.active
    values = [value[0] for value in ws.iter_rows(min_row=2, min_col=3, max_col=3, values_only=True)]

    return values

# excel = load_workbook(f'./')

if __name__ == '__main__':
    print(excel_iterator())