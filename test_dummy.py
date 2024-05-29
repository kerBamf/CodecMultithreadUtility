import os
import time
import math
from openpyxl import load_workbook
from datetime import datetime
import xml.etree.ElementTree as ET

from chain_update import check_codec

xml_string = '<?xml version="1.0"?><Command><SystemUnitSoftwareUpgradeResult status="OK"/></Command>'

xml_root = ET.fromstring(xml_string)

xcel = load_workbook(f'./codec_lists/{input("Please input codec file: ")}')
xcel_active = xcel.active
codec_ips = xcel_active.iter_rows(min_row=2, min_col=3, max_col=3, values_only=True)

for value in codec_ips:
  print(value[0])

if __name__ == '__main__':
  # print(xml_root[0].attrib['status'])
  # codec_info = check_codec('172.16.131.163')
  print(codec_ips)
