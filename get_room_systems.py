from os import environ
from os import listdir, getcwd
from os.path import isfile, join
from dotenv import load_dotenv
import xml.etree.ElementTree as ET
from Utils.logger import log_info
from Utils.cod_get import cod_get
from Utils.cod_post import cod_post, cod_session_start, cod_session_end
from Utils.excel_parser import excel_parser
import openpyxl
import concurrent.futures
from dataclasses import dataclass, astuple


# **************
# Loading dependencies and file retrieval functionality
# **************

 
# Loading environment variable
load_dotenv()
PASSCODE = environ.get("PASSCODE")
LOGPATH = environ.get("LOGPATH")
headers = {"Authorization": f"basic {PASSCODE}"}


# Setting up logger
def message(string, ip, path=LOGPATH):
    log_info(string, ip, path)
    print(string)


# Setting Up File Retrieval
def find_path():
    path = f"{getcwd()}/../excel_files"
    return path


def user_choice(files):
    choice = int(
        input(
            f'Select which excel file to use:\r\n{files["options"]}\r\nFile number selection: '
        )
    )
    for key in list(files["files"].keys()):
        if choice == key:
            if "y" == input(
                f'\n\rYou have selected {files["files"][key]} Proceed? (y/n): '
            ):
                return choice
            else:
                print("Selection Cancelled")
                return user_choice()
    print("Your selection is invalid.")
    return user_choice()


def select_file(path):
    file_list = [file for file in listdir(path) if isfile(join(path, file))]
    list_dict = {}
    option_string = ""
    for idx, file in enumerate(file_list):
        list_dict.update({idx: file})
        option_string = option_string + f"\t{file}: {idx}\r\n"

    output = {"files": list_dict, "options": option_string}

    selected_file = output["files"][user_choice(output)]

    return selected_file


# **************************
# Retrieving Info from Codec
# **************************


def startSession(codec):
    try:
        return cod_session_start(codec.ip)
    except Exception as error:
        print(f"{codec.name} Error: Error starting session - {error.text}")
        raise


def check_cams(codec, cookie):
    try:
        response = cod_get(codec.ip, "Status/Cameras/Camera/Model", cookie)
        root = ET.fromstring(response)
        elements = root.findall(".//Model")
        camString = ""
        for element in elements:
            text = element.text
            if camString == "":
                camString += f"{text};"
            else:
                camString += f" {text};"
        return camString
    except Exception as err:
        print(f"{codec.name} Error: Unable to check cams - {err.text}")
        raise


def check_inputs(codec, cookie):
    try:
        response = cod_get(codec.ip, "Configuration/Video/Input/Connector", cookie)
        root = ET.fromstring(response)
        raw_inputs = root.findall(".//Name")
        inputString = ""
        for input in raw_inputs:
            if (
                input.text != None
                and input.text != ""
                and type(input.text) == str
                and input.text.find("Camera") == -1
            ):
                if inputString == "":
                    inputString += f"{input.text};"
                else:
                    inputString += f" {input.text};"
        return inputString
    except Exception as err:
        print(f"{codec.name} Error: Unable to check cams - {err.text}")
        raise


def check_lwr_macro(codec, cookie):
    try:
        macro_xml = f"""<Command>
    <Macros>
        <Macro>
            <Get></Get>
        </Macro>
    </Macros>
</Command>"""

        response = cod_post(codec.ip, macro_xml, cookie)
        if response.find("Lightware Integration") != -1:
            return True
        else:
            return False
    except Exception as err:
        print(f"{codec.name} Error: Unable to check macros - {err.text}")
        raise


lwr_src_XML = f"""<Command>
    <UserInterface>
        <Presentation>
            <ExternalSource>
                <List></List>
            </ExternalSource>
        </Presentation>
    </UserInterface>
</Command>"""


def check_lwr_sources(codec, cookie):
    try:
        response = cod_post(codec.ip, lwr_src_XML, cookie)
        root = ET.fromstring(response)
        elements = root.findall(".//Name")
        inputString = ""
        for element in elements:
            text = element.text
            if text != "":
                if inputString == "":
                    inputString += f"{text};"
                else:
                    inputString += f" {text};"
        return inputString
    except Exception as error:
        message(f"Unable to retrieve info at {codec.name} --> {error.text}", codec.name)
        raise


def check_outputs(codec, cookie):
    def getElements(response, attr):
        root = ET.fromstring(response)
        elements = root.findall(f".//{attr}")
        return elements

    try:
        conn_res = cod_get(codec.ip, "Status/Video/Output/Connector/Connected", cookie)
        conn = getElements(conn_res, "Connected")
        outputString = ""
        for idx, port in enumerate(conn):
            text = port.text
            if text == "True":
                name_res = cod_get(
                    codec.ip,
                    f"Status/Video/Output/Connector[{idx+1}]/ConnectedDevice/Name",
                    cookie,
                )
                name = getElements(name_res, "Name")[0].text

                size_res = cod_get(
                    codec.ip,
                    f"Status/Video/Output/Connector[{idx+1}]/ConnectedDevice/ScreenSize",
                    cookie,
                )
                size = getElements(size_res, "ScreenSize")[0].text
                output = f"{name} {size}"
                if outputString == "":
                    outputString += output + ";"
                else:
                    outputString += f" {output};"
        return outputString

    except Exception as err:
        print(f"{codec.name} Error: Unable to check outputs - {err.text}")
        raise


# Calling functions to fully check system
def check_system(codec):
    if codec.status == "No Response":
        codec.error == "Codec offline. Unable to gather information"
        return codec

    cookie = None

    try:
        cookie = startSession(codec)
        cams = check_cams(codec, cookie)
        codec.cameras = cams
        vid_inputs = check_inputs(codec, cookie)
        codec.input_sources = vid_inputs
        if vid_inputs.find("Source") != -1 and type(vid_inputs) == str:
            codec.lightware = "Present"
            lwr_sources = check_lwr_sources(codec, cookie)
            codec.lwr_input_sources = lwr_sources
        if (
            codec.type_description != "Cisco Room 55"
            and codec.type_description != "Cisco Desk Pro"
        ):
            outputs = check_outputs(codec, cookie)
            codec.output_devices = outputs
        # print(
        #     f"{codec.name} {codec.type_description} {codec.cameras} {codec.input_sources} {codec.lightware} {codec.lwr_input_sources} {codec.output_devices}"
        # )
        cod_session_end(codec.ip, cookie)
        return codec

    except Exception as err:
        codec.error = err.text
        if (
            cookie
            and type(cookie) == "<class 'str'>"
            and cookie.find("SessionId") != -1
        ):
            cod_session_end(codec.ip, cookie)
        return codec

    # return codec


# ***********************
# Main function
# ***********************


def get_room_systems():
    path = find_path()
    filepath = f"{path}/{select_file(path)}"

    # Loading excel workbook and converting to list of Codec classes
    excel_import = openpyxl.load_workbook(filepath)
    worksheet = excel_import.active

    @dataclass
    class Codec:
        name: str = ""
        ip: str = ""
        status: str = ""
        serial: str = ""
        type_description: str = ""
        error: str = ""
        cameras: str = ""
        input_sources: str = ""
        lightware: str = ""
        lwr_input_sources: str = ""
        output_devices: str = ""

    codec_list = []

    for value in worksheet.iter_rows(min_row=2, min_col=2, max_col=6, values_only=True):
        codec = Codec(
            name=value[0],
            ip=value[1],
            status=value[2],
            serial=value[3],
            type_description=value[4],
        )
        codec_list.append(codec)

    codecs_processed = [
        [
            "System Name",
            "IP Address",
            "Status",
            "Hardware Serial Number",
            "Specific System Type Description",
            "Error",
            "Cameras",
            "Input Sources",
            "Lightware",
            "Lightware Input Sources",
            "Output Devices",
        ]
    ]

    # Feeding through multithreader to obtain data from codecs
    with concurrent.futures.ThreadPoolExecutor() as executor:
        futures = {executor.submit(check_system, codec): codec for codec in codec_list}

    for future in concurrent.futures.as_completed(futures):
        codec = future.result()
        print(list(astuple(codec)))
        codec_values = list(astuple(codec))
        codecs_processed.append(codec_values)

    # Creating new Excel file
    new_wb = openpyxl.Workbook()
    new_ws = new_wb.active
    new_ws.title = "RoomSystems"
    ft = openpyxl.styles.Font(bold=True)

    for row in codecs_processed:
        new_ws.append(row)

    for row in new_ws["A1:K1"]:
        for cell in row:
            cell.font = ft

    new_wb.save("../output_files/FullSystemOutput.xlsx")


if __name__ == "__main__":
    get_room_systems()
