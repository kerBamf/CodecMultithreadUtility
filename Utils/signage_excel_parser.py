from os import listdir, getcwd
from os.path import isfile, join
from openpyxl import load_workbook
from dataclasses import dataclass


class Custom_Exception(Exception):
    pass


def find_path():
    path = f"{getcwd()}/../excel_files"
    return path


# Building file list
def select_file(path):
    file_list = [file for file in listdir(path) if isfile(join(path, file))]
    list_dict = {}
    option_string = ""
    for idx, file in enumerate(file_list):
        list_dict.update({idx: file})
        option_string = option_string + f"\t{file}: {idx}\r\n"

    output = {"files": list_dict, "options": option_string}

    selected_file = output["files"][user_choice(output)]

    return selected_file


# User selection functionality
def user_choice(files):
    choice = int(
        input(
            f'Select which excel file to use:\r\n{files["options"]}\r\nFile number selection: '
        )
    )
    for key in list(files["files"].keys()):
        if choice == key:
            if "y" == input(
                f'\n\rYou have selected {files["files"][key]} Proceed? (y/n): '
            ):
                return choice
            else:
                print("Selection Cancelled")
                return user_choice()
    print("Your selection is invalid.")
    return user_choice()


# Setting up class for codecs
@dataclass
class Codec:
    name: str
    ip: str
    url: str


# Main function
def signage_excel_parser(path="", file=""):
    if path == "":
        path = find_path()
    if file == "":
        file = select_file(find_path())
    wb = load_workbook(f"{path}/{file}", data_only=True)
    ws = wb.active
    codecs = []
    for cell in ws.iter_rows(min_row=2, min_col=2, max_col=4, values_only=True):
        codec = Codec(cell[0], cell[1], cell[2])
        codecs.append(codec)
    return codecs


# excel = load_workbook(f'./')

if __name__ == "__main__":
    # Finding relative path
    print(signage_excel_parser())
